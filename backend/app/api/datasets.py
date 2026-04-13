from fastapi import APIRouter, UploadFile, Form, HTTPException, File, Depends
from sqlalchemy.orm import Session
from typing import Annotated
from app.core.permissions import check_permissions
from sqlalchemy import text
from typing import Optional, List, Any
import pandas as pd
from io import BytesIO
from pydantic import BaseModel
import re
import datetime
from openpyxl import load_workbook
from app.core.database import get_db, engine
from app.models.dataset import Dataset
from app.models.dataset_column import DatasetColumn
from app.models.dataset_row import DatasetRow
from app.utils.type_inference import infer_column_type
from fastapi.responses import StreamingResponse
from fastapi import Query
import json
from collections import OrderedDict
import threading

# 🔹 SIMPLE LRU CACHE FOR DASHBOARD DATA
class DatasetCache:
    def __init__(self, capacity=256):
        self.cache = OrderedDict()
        self.capacity = capacity
        self.lock = threading.Lock()

    def get(self, key):
        with self.lock:
            if key not in self.cache:
                return None
            self.cache.move_to_end(key)
            return self.cache[key]

    def set(self, key, value):
        with self.lock:
            if key in self.cache:
                self.cache.move_to_end(key)
            self.cache[key] = value
            if len(self.cache) > self.capacity:
                self.cache.popitem(last=False)

    def invalidate(self, dataset_id: int):
        with self.lock:
            # Keys are tuples (dataset_id, function_name, *args)
            keys_to_remove = [k for k in self.cache.keys() if k[0] == dataset_id]
            for k in keys_to_remove:
                del self.cache[k]
            print(f"Invalidated cache for dataset_id: {dataset_id}")

global_dataset_cache = DatasetCache()

router = APIRouter(prefix="/datasets", tags=["Datasets"])

def get_dataset_df(dataset: Dataset, db: Session) -> pd.DataFrame:
    """Helper to get DataFrame from either dynamic table or legacy DatasetRow"""
    if dataset.table_name:
        try:
            with engine.begin() as conn:
                return pd.read_sql_table(dataset.table_name, conn)
        except Exception as e:
            print(f"Error reading table {dataset.table_name}: {e}")
            return pd.DataFrame()
    else:
        rows = db.query(DatasetRow).filter_by(dataset_id=dataset.id).all()
        return pd.DataFrame([r.row_data for r in rows])

@router.get("/", dependencies=[Depends(check_permissions("view_tracker"))])
def list_datasets(db: Annotated[Session, Depends(get_db)]):
    datasets = db.query(Dataset).order_by(Dataset.id.desc()).all()
    return [
        {
            "id": d.id,
            "project": d.project,
            "department": d.department,
            "employeeName": d.uploaded_by,
            "fileName": d.name,
            "uploadedBy": d.uploaded_by,
            "uploadDate": d.created_at.strftime("%Y-%m-%d") if d.created_at else None,
            "fileType": d.file_type,
            "records": d.row_count,
            "status": "Completed",
            "uploadDateISO": d.created_at.isoformat() if d.created_at else None
        }
        for d in datasets
    ]

# 🔹 PREVIEW DATA (EYE BUTTON)
@router.get("/{dataset_id}/excel-view", dependencies=[Depends(check_permissions("view_tracker"))])
def get_excel_view(dataset_id: int, db: Annotated[Session, Depends(get_db)]):
    # Check cache first
    cache_key = (dataset_id, "excel_view")
    cached_data = global_dataset_cache.get(cache_key)
    if cached_data:
        return cached_data

    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    columns = (
        db.query(DatasetColumn)
        .filter(DatasetColumn.dataset_id == dataset_id)
        .order_by(DatasetColumn.id)
        .all()
    )

    headers = [c.column_name for c in columns]
    if dataset.table_name:
        try:
            with engine.begin() as conn:
                df = pd.read_sql_query(text(f'SELECT * FROM "{dataset.table_name}" LIMIT 1000'), conn)
        except Exception as e:
            print(f"Error reading table {dataset.table_name}: {e}")
            df = pd.DataFrame()
    else:
        rows = db.query(DatasetRow).filter_by(dataset_id=dataset_id).limit(1000).all()
        if rows:
            df = pd.DataFrame([r.row_data for r in rows])
        else:
            df = pd.DataFrame()
    
    if not df.empty:
        # If headers logic is out of sync, trust the DF
        if not headers:
            headers = df.columns.tolist()
        
        # Ensure we only try to get columns that exist in DF
        valid_headers = [h for h in headers if h in df.columns]
        # Add any new columns in DF not in headers
        extra_cols = [c for c in df.columns if c not in valid_headers]
        final_headers = valid_headers + extra_cols
        
        df = df.fillna("")
        data = df[final_headers].values.tolist()
        headers = final_headers
    else:
        data = []

    result = {
        "id": dataset.id,
        "name": dataset.name,
        "type": dataset.file_type,
        "size": dataset.row_count,
        "date": dataset.created_at.strftime("%Y-%m-%d"),
        "uploadedBy": dataset.uploaded_by or "System",
        "fileData": {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": headers,
                    "data": data
                }
            ]
        }
    }

    global_dataset_cache.set(cache_key, result)
    return result


# 🔹 DOWNLOAD DATA AGAIN
@router.get("/{dataset_id}/download", dependencies=[Depends(check_permissions("view_tracker"))])
def download_dataset(
    dataset_id: int, 
    format: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    dataset = db.query(Dataset).filter_by(id=dataset_id).first()
    if not dataset:
        return {"error": "Not found"}

    df = get_dataset_df(dataset, db)

    # Determine export format: use query param if provided, else use original file type
    export_format = format.lower() if format else dataset.file_type.lower()
    
    stream = BytesIO()
    
    if export_format == "csv":
        df.to_csv(stream, index=False)
        media_type = "text/csv"
        filename = f"{dataset.name.rsplit('.', 1)[0]}.csv"
    else:
        # Default to Excel for xlsx, xls, or any other format
        df.to_excel(stream, index=False)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{dataset.name.rsplit('.', 1)[0]}.xlsx"

    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# 🔹 CHART DATA API
@router.get("/{dataset_id}/chart")
def get_chart_data(
    dataset_id: int,
    x: str = Query(...),
    y: str = Query(...),
    db: Session = Depends(get_db),
):
    # Check cache first
    cache_key = (dataset_id, "chart", x, y)
    cached_data = global_dataset_cache.get(cache_key)
    if cached_data:
        return cached_data

    dataset = db.query(Dataset).filter_by(id=dataset_id).first()
    if not dataset:
        return {"x": [], "y": [], "count": 0}

    x_vals = []
    y_vals = []
    
    if dataset.table_name:
        try:
            with engine.begin() as conn:
                # Need to quote the column names in case they have spaces
                query = text(f'SELECT "{x}", "{y}" FROM "{dataset.table_name}"')
                df = pd.read_sql_query(query, conn)
        except Exception as e:
            print(f"Error fetching chart data: {e}")
            df = pd.DataFrame()
    else:
        df = get_dataset_df(dataset, db)

    if not df.empty and x in df.columns and y in df.columns:
        # Filter for valid numeric Y values
        df_valid = df[pd.to_numeric(df[y], errors='coerce').notna()]
        x_vals = df_valid[x].tolist()
        y_vals = df_valid[y].astype(float).tolist()

    result = {
        "x": x_vals,
        "y": y_vals,
        "count": len(x_vals)
    }
    global_dataset_cache.set(cache_key, result)
    return result

class UpdateDatasetRequest(BaseModel):
    headers: List[str]
    data: List[Any]

@router.put("/{dataset_id}/data")
def update_dataset_data(
    dataset_id: int,
    payload: UpdateDatasetRequest,
    db: Session = Depends(get_db)
):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    # Update columns
    # First, delete existing columns metadata
    db.query(DatasetColumn).filter(DatasetColumn.dataset_id == dataset_id).delete()
    
    # Add new columns metadata
    for col_name in payload.headers:
        db.add(DatasetColumn(
            dataset_id=dataset_id,
            column_name=col_name,
            data_type="string" 
        ))

    # Update rows
    if dataset.table_name:
        # Update dynamic table
        try:
            # Check if data is already list of dicts or list of lists
            if payload.data and isinstance(payload.data[0], dict):
                new_df = pd.DataFrame(payload.data)
                # Reorder to match headers and handle missing columns in some rows
                cols_to_use = [h for h in payload.headers if h in new_df.columns]
                new_df = new_df[cols_to_use]
            else:
                new_df = pd.DataFrame(payload.data, columns=payload.headers)
            
            new_df.to_sql(dataset.table_name, engine, if_exists='replace', index=False)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to update table: {e}")
    else:
        # Legacy update
        db.query(DatasetRow).filter(DatasetRow.dataset_id == dataset_id).delete()
        new_rows = []
        for row_entry in payload.data:
            if isinstance(row_entry, dict):
                row_dict = row_entry
            else:
                row_dict = {}
                for i, val in enumerate(row_entry):
                    if i < len(payload.headers):
                        row_dict[payload.headers[i]] = val
            
            new_rows.append(DatasetRow(
                dataset_id=dataset_id,
                row_data=row_dict
            ))
        db.bulk_save_objects(new_rows)
    
    # Update row count
    dataset.row_count = len(payload.data)
    
    # Invalidate Cache
    global_dataset_cache.invalidate(dataset_id)
    
    return {"message": "Dataset updated successfully"}

def process_excel(file_bytes: bytes):
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = wb.active

    # 1. Handle merged cells
    merged_data = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = sheet.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_data[(row, col)] = top_left_val

    # 2. Extract all non-empty cells to find the bounding box
    cells_dict = {}
    min_r, max_r, min_c, max_c = float('inf'), -1, float('inf'), -1
    
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            val = merged_data.get((row, col), cell.value)
            
            if val is not None and str(val).strip() != "":
                cells_dict[(row, col)] = val
                min_r = min(min_r, row)
                max_r = max(max_r, row)
                min_c = min(min_c, col)
                max_c = max(max_c, col)

    if not cells_dict:
        return []

    # 3. Robust Header Row Detection
    header_start_row = None
    header_depth_detected = 1
    final_headers = []
    
    for r in range(min_r, min(min_r + 20, max_r + 1)):
        non_null_cells = 0
        text_cells = 0
        numeric_cells = 0
        
        for c in range(min_c, max_c + 1):
            val = cells_dict.get((r, c))
            if val is not None and str(val).strip() != "":
                non_null_cells += 1
                if isinstance(val, (int, float, complex)) and not isinstance(val, bool):
                    numeric_cells += 1
                else:
                    text_cells += 1
                    
        min_required = min(5, max_c - min_c + 1)
        if non_null_cells >= min_required:
            text_ratio = text_cells / non_null_cells
            numeric_ratio = numeric_cells / non_null_cells
            
            if text_ratio >= 0.70 and numeric_ratio <= 0.30:
                best_d = 1
                max_score = -1
                best_headers = []
                
                for d in range(1, min(5, max_r - r + 2)):
                    headers_at_d = []
                    header_block = {}
                    
                    for hr in range(r, r + d):
                        last_val = ""
                        for c in range(min_c, max_c + 1):
                            val = cells_dict.get((hr, c))
                            if val is not None and str(val).strip() != "":
                                last_val = str(val).strip()
                            
                            if hr < r + d - 1:
                                header_block[(hr, c)] = last_val
                            else:
                                header_block[(hr, c)] = str(val).strip() if val is not None else ""
                                
                    for c in range(min_c, max_c + 1):
                        col_parts = []
                        for hr in range(r, r + d):
                            part = header_block.get((hr, c), "")
                            if part and (not col_parts or col_parts[-1] != part):
                                col_parts.append(part)
                                
                        header_name = " ".join(col_parts).lower().strip()
                        header_name = re.sub(r'[.\-]', ' ', header_name)
                        header_name = re.sub(r'[^a-z0-9_\s]', '', header_name)
                        header_name = re.sub(r'\s+', '_', header_name).strip('_')
                        
                        if not header_name:
                            header_name = f"column_{c}"
                            
                        headers_at_d.append(header_name)
                        
                    unique_count = len(set(headers_at_d))
                    generic_count = sum(1 for h in headers_at_d if h.startswith("column_"))
                    
                    score = unique_count - (generic_count * 0.5)
                    
                    if score > max_score:
                        max_score = score
                        best_d = d
                        best_headers = headers_at_d
                
                has_generic = any(h.startswith("column_") for h in best_headers)
                
                if not has_generic:
                    header_start_row = r
                    header_depth_detected = best_d
                    final_headers = best_headers
                    break

    if header_start_row is None:
        raise ValueError("Could not detect a valid header row. Ensure your sheet has a header with at least 5 text columns and no generic placeholders.")
        
    header_rows_end = header_start_row + header_depth_detected - 1

    unique_headers = []
    header_counts = {}
    for h in final_headers:
        if h in header_counts:
            header_counts[h] += 1
            unique_headers.append(f"{h}_{header_counts[h]}")
        else:
            header_counts[h] = 0
            unique_headers.append(h)

    # 5. Extract rows
    data = []
    for r in range(header_rows_end + 1, max_r + 1):
        row_data = {}
        is_empty = True
        for idx, c in enumerate(range(min_c, max_c + 1)):
            val = cells_dict.get((r, c))
            
            if isinstance(val, (datetime.date, datetime.datetime)):
                val = val.isoformat()
                
            if val is not None:
                is_empty = False
                
            row_data[unique_headers[idx]] = val
            
        if not is_empty:
            data.append(row_data)

    return data

@router.post("/upload", dependencies=[Depends(check_permissions("upload_tracker"))])
async def upload_dataset(
    file: UploadFile = File(...),
    industry: Optional[str] = Form(None),
    project: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    employeeName: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    Accepts CSV or Excel file and stores dataset, columns, and rows in DB.
    Returns dataset metadata compatible with frontend.
    """

    # --- PROJECT INTEGRATION ---
    from app.crud import project as crud_project
    from app.schemas.project import ProjectCreate

    existing_project = crud_project.get_project_by_name(db, project)
    if not existing_project:
        # Create new project if it doesn't exist
        # Generate a project_id slug
        project_id_slug = re.sub(r'[^a-zA-Z0-9-]', '', project.lower().replace(' ', '-'))
        # Ensure it's not empty and add a simple suffix if needed
        if not project_id_slug:
            project_id_slug = f"proj-{datetime.datetime.now().strftime('%y%m%d%H%M')}"
        
        new_project_data = ProjectCreate(
            project_id=project_id_slug,
            name=project,
            manager=employeeName or "Unassigned",
            status="Planning",
            budget=0.0,
            teamSize=0,
            employee_name=employeeName
        )
        crud_project.create_project(db, new_project_data)
    # ---------------------------

    # 1️⃣ Read file into DataFrame
    try:
        contents = await file.read()
        if file.filename.endswith(".csv"):
            df = pd.read_csv(BytesIO(contents))
            df = df.fillna("")
        else:
            processed_data = process_excel(contents)
            if not processed_data:
                raise HTTPException(status_code=400, detail="No readable data found.")
            df = pd.DataFrame(processed_data)
            df = df.fillna("")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")

    # 1.5 Construct new filename: ProjectName_FileName (without extension)
    # Aggressively strip multiple extensions if present
    base_filename = file.filename
    while '.' in base_filename:
        base_filename = base_filename.rsplit('.', 1)[0]
        
    new_filename = f"{project}_{base_filename}" if project else base_filename
    file_extension = file.filename.split(".")[-1].upper()

    # 0️⃣ Check for duplicates with the new filename
    # "one department cannot upload duplicate tracker(excel)"
    if department:
        existing = db.query(Dataset).filter(
            Dataset.department == department, 
            Dataset.name == new_filename
        ).first()
        if existing:
            raise HTTPException(
                status_code=400, 
                detail=f"Dataset '{new_filename}' already uploaded for department '{department}'."
            )

    # 2️⃣ Store dataset metadata
    dataset = Dataset(
        name=new_filename,
        industry=industry,
        project=project,
        department=department,
        uploaded_by=employeeName,
        file_type=file_extension,
        row_count=len(df)
    )
    db.add(dataset)
    db.commit()
    db.refresh(dataset)

    # 3️⃣ Create Dynamic Table and Insert Data
    sanitized_project = re.sub(r'[^a-zA-Z0-9_]', '_', project).lower() if project else ""
    sanitized_file = re.sub(r'[^a-zA-Z0-9_]', '_', base_filename).lower()
    
    # User requested: project name_filename at the start, id at the end for technical uniqueness
    table_name_base = f"{sanitized_project}_{sanitized_file}" if sanitized_project else sanitized_file
    table_name = f"{table_name_base}_{dataset.id}"[:63] # Postgres limit 63 chars

    try:
        # Create table and insert data
        df.to_sql(table_name, engine, if_exists='fail', index=False)
        
        # Update dataset with table_name
        dataset.table_name = table_name
        db.commit()
    except Exception as e:
        # Rollback metadata if table creation fails
        db.delete(dataset)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed to create table: {e}")

    # 4️⃣ Store column metadata (Keeping for consistency/schema endpoint)
    for col in df.columns:
        db.add(DatasetColumn(
            dataset_id=dataset.id,
            column_name=col,
            data_type=infer_column_type(df[col])
        ))
    db.commit()

    # Invalidate cache if overwriting an existing dataset
    global_dataset_cache.invalidate(dataset.id)

    # 5️⃣ Return metadata for frontend tracker
    return {
        "id": dataset.id,
        "project": dataset.project,
        "department": dataset.department,
        "employeeName": dataset.uploaded_by,
        "fileName": dataset.name,
        "uploadedBy": dataset.uploaded_by,
        "uploadDate": dataset.created_at.strftime("%Y-%m-%d"),
        "fileType": dataset.file_type,
        "records": dataset.row_count,
        "status": "Completed",
        "uploadDateISO": dataset.created_at.isoformat()
    }

# ✅ THIS IS WHERE YOUR QUESTIONED CODE GOES
@router.get("/{dataset_id}/schema")
def get_schema(dataset_id: int, db: Session = Depends(get_db)):
    return db.query(DatasetColumn).filter_by(dataset_id=dataset_id).all()


@router.get("/{dataset_id}/data")
def get_data(dataset_id: int, db: Session = Depends(get_db)):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        return []
    
    # For dynamic tables, query only the first 1000 rows from the database directly
    if dataset.table_name:
        try:
            with engine.begin() as conn:
                query = text(f'SELECT * FROM "{dataset.table_name}" LIMIT 1000')
                df = pd.read_sql_query(query, conn)
                return df.fillna("").to_dict(orient='records')
        except Exception as e:
            print(f"Error fetching data: {e}")
            return []

    # Fallback for legacy datasets
    rows = db.query(DatasetRow).filter_by(dataset_id=dataset_id).limit(1000).all()
    if rows:
        df = pd.DataFrame([r.row_data for r in rows])
        return df.fillna("").to_dict(orient='records')
    return []

class UpdateDatasetMetadataRequest(BaseModel):
    project: Optional[str] = None
    department: Optional[str] = None
    employeeName: Optional[str] = None

@router.put("/{dataset_id}")
def update_dataset_metadata(
    dataset_id: int,
    payload: UpdateDatasetMetadataRequest,
    db: Session = Depends(get_db)
):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    if payload.project is not None:
        dataset.project = payload.project
    if payload.department is not None:
        dataset.department = payload.department
    if payload.employeeName is not None:
        dataset.uploaded_by = payload.employeeName
    
    db.commit()
    db.refresh(dataset)
    
    # Invalidate cache
    global_dataset_cache.invalidate(dataset_id)
    
    return {"message": "Dataset metadata updated successfully"}

@router.delete("/{dataset_id}", dependencies=[Depends(check_permissions("delete_tracker"))])
def delete_dataset(dataset_id: int, db: Annotated[Session, Depends(get_db)]):
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()

    if not dataset:
        return {"error": "Dataset not found"}

    # Drop dynamic table if exists
    if dataset.table_name:
        try:
            # Use text() for raw SQL to drop table
            db.execute(text(f'DROP TABLE IF EXISTS "{dataset.table_name}"'))
        except Exception as e:
            print(f"Error dropping table {dataset.table_name}: {e}")

    # Delete child rows first (FK safety) - for legacy data
    db.query(DatasetRow).filter(DatasetRow.dataset_id == dataset_id).delete()
    db.query(DatasetColumn).filter(DatasetColumn.dataset_id == dataset_id).delete()

    # Invalidate cache
    global_dataset_cache.invalidate(dataset_id)

    db.delete(dataset)
    db.commit()

    return {"message": "Dataset deleted successfully"}

class ProcessDatasetRequest(BaseModel):
    row_indices: Optional[List[int]] = None

@router.post("/{dataset_id}/process")
def process_dataset_data(
    dataset_id: int,
    payload: ProcessDatasetRequest = None,
    db: Session = Depends(get_db)
):
    """
    Triggers re-processing of dataset data:
    1. Re-infers column types
    2. Transforms data to normalized formats (e.g. dates to ISO)
    3. Updates column metadata and underlying table/rows
    Supports partial processing via payload.row_indices
    """
    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")

    df = get_dataset_df(dataset, db)
    if df.empty:
        raise HTTPException(status_code=400, detail="No data available to process")

    # 1. Clean and Transform Data
    new_columns_metadata = []
    
    # Identify explicit subset to process
    subset_indices = payload.row_indices if payload and payload.row_indices else list(range(len(df)))
    target_df = df.iloc[subset_indices].copy()
    
    for col in df.columns:
        # Preprocessing: Clean up string values only on target_df
        if target_df[col].dtype == 'object':
            target_df[col] = target_df[col].apply(lambda x: pd.NA if pd.isna(x) or str(x).strip().lower() in ['nan', 'none', '', 'null'] else str(x).strip())
            target_df[col] = target_df[col].apply(lambda x: x.replace('--', '-') if isinstance(x, str) else x)
        
        non_null_mask = target_df[col].notna()
        non_null_count = non_null_mask.sum()
        
        if non_null_count == 0:
            new_columns_metadata.append({"name": col, "type": "string"})
            continue
            
        temp_col = target_df[col].apply(lambda x: x.replace(',', '') if isinstance(x, str) else x)
        numeric_series = pd.to_numeric(temp_col, errors='coerce')
        valid_numeric_count = numeric_series.notna().sum()
        
        inferred_type = None

        # Robust Date Parsing: Try default first (which handles ISO perfectly)
        # Then try dayfirst=True (which handles DD/MM well, but breaks some ISO strings)
        # Use whichever yields more valid dates
        date_series_default = pd.to_datetime(target_df[col], errors='coerce')
        valid_date_count_default = date_series_default.notna().sum()
        
        date_series_df = pd.to_datetime(target_df[col], errors='coerce', dayfirst=True)
        valid_date_count_df = date_series_df.notna().sum()
        
        if valid_date_count_df > valid_date_count_default:
            date_series = date_series_df
            valid_date_count = valid_date_count_df
        else:
            date_series = date_series_default
            valid_date_count = valid_date_count_default

        num_ratio = valid_numeric_count / non_null_count
        date_ratio = valid_date_count / non_null_count

        # If it's mostly numeric, and doesn't look heavily like dates, check for integers
        # Only treat as date if date_ratio > num_ratio, OR date_ratio >= 0.4 and it's heavily formatted strings (like '2026-07-31' where numeric fails)
        if date_ratio >= 0.4 and date_ratio > num_ratio:
            formatted_dates = date_series.dt.strftime('%Y-%m-%d')
            target_df[col] = formatted_dates.where(date_series.notna(), target_df[col])
            inferred_type = "date"
        elif num_ratio >= 0.7:
            numeric_vals = numeric_series.dropna()
            if all(val == float(int(val)) for val in numeric_vals):
                target_df[col] = numeric_series.round().astype('Int64')
                inferred_type = "integer"
            else:
                target_df[col] = numeric_series
                inferred_type = "float"
        elif date_ratio >= 0.4:
            formatted_dates = date_series.dt.strftime('%Y-%m-%d')
            target_df[col] = formatted_dates.where(date_series.notna(), target_df[col])
            inferred_type = "date"
        else:
            inferred_type = "string"

        if inferred_type is None:
            inferred_type = infer_column_type(target_df[col])
            
        # Apply the transformed subset back to the main dataframe
        df.iloc[subset_indices, df.columns.get_loc(col)] = target_df[col]
            
        new_columns_metadata.append({
            "name": col,
            "type": inferred_type
        })

    # 2. Persist Optimized Data
    if dataset.table_name:
        try:
            df.to_sql(dataset.table_name, engine, if_exists='replace', index=False)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to update table during optimize: {e}")
    else:
        # Legacy update
        db.query(DatasetRow).filter(DatasetRow.dataset_id == dataset_id).delete()
        new_rows = []
        for _, row in df.iterrows():
            # Handle NaN values for JSON serialization
            processed_row = row.replace({pd.NA: None, float('nan'): None}).to_dict()
            new_rows.append(DatasetRow(
                dataset_id=dataset_id,
                row_data=processed_row
            ))
        db.bulk_save_objects(new_rows)

    # 3. Update Metadata
    db.query(DatasetColumn).filter(DatasetColumn.dataset_id == dataset_id).delete()
    for col_info in new_columns_metadata:
        db.add(DatasetColumn(
            dataset_id=dataset_id,
            column_name=col_info["name"],
            data_type=col_info["type"]
        ))
    
    db.commit()
    
    # Invalidate cache
    global_dataset_cache.invalidate(dataset_id)
    
    return {
        "message": "Dataset optimized and normalized successfully",
        "columns": new_columns_metadata,
        "rowCount": len(df)
    }

